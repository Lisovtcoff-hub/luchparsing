
"""Модуль export_xlsx.

Содержит прикладную логику и точки входа проекта.
"""

from __future__ import annotations

import json
import ast
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from database.db import Database


HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
BORDER_ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

FILL_RED = PatternFill(start_color="FFFFE5E5", end_color="FFFFE5E5", fill_type="solid")
FILL_GREEN = PatternFill(start_color="FFE6F4EA", end_color="FFE6F4EA", fill_type="solid")
FILL_GRAY = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")

HYPERLINK_FONT = Font(color="FF0563C1", underline="single")

INVALID_SHEET_CHARS = set(r"[]:*?\/")


def _safe_title(s: str) -> str:
    """Функция _safe_title.

    Параметры:
        s: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    s = "".join((" " if ch in INVALID_SHEET_CHARS else ch) for ch in (s or ""))
    s = " ".join(s.split())
    return s[:31] or "Лист"


def _unique_sheet_title(wb: Workbook, base: str) -> str:
    """Функция _unique_sheet_title.

    Параметры:
        wb: Описание параметра.
        base: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    name = _safe_title(base)
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        suffix = f" {i}"
        candidate = (name[: (31 - len(suffix))] + suffix)
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _fmt_preset_header(preset: dict) -> str:
    """Функция _fmt_preset_header.

    Параметры:
        preset: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    places = preset["places"]
    weight = preset["weight_kg"]
    vol = round(float(preset["volume_m3"]), 3)
    weight_str = int(weight) if float(weight).is_integer() else weight
    s = f"{weight_str}".replace(".", ",")
    v = f"{vol}".replace(".", ",")
    return f"{s} кг ({places} мест, {v} куба)"


def _autosize(ws: Worksheet) -> None:
    """Функция _autosize.

    Автоширина подбирается по содержимому, но для отдельных колонок задаются
    более узкие ограничения (например, для цены и дней).

    Параметры:
        ws: Описание параметра.
    """
    for col in ws.columns:
        max_len = 1
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)

        header_val = ws.cell(row=1, column=col_idx).value
        header = "" if header_val is None else str(header_val)

        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if "\n" in val:
                max_len = max(max_len, max(len(line) for line in val.splitlines()))
            else:
                max_len = max(max_len, len(val))

        min_w, max_w, pad = 12, 52, 2

        if "Цена" in header:
            min_w, max_w, pad = 8, 18, 1

        if header.strip() == "Дни":
            min_w, max_w, pad = 6, 10, 1

        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + pad), max_w)


def _border_with(
    base: Border,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    """Функция _border_with.

    Параметры:
        base: Описание параметра.
        left: Описание параметра.
        right: Описание параметра.
        top: Описание параметра.
        bottom: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    return Border(
        left=left if left is not None else base.left,
        right=right if right is not None else base.right,
        top=top if top is not None else base.top,
        bottom=bottom if bottom is not None else base.bottom,
    )


def _apply_table_styles(ws: Worksheet, rows_count: int, cols_count: int, presets_count: int) -> None:
    """
    Стили:
    - тонкие границы по всем ячейкам
    - жирная шапка
    - жирная рамка по периметру всей таблицы
    - жирная рамка блока A:C (Компания/Дни/Тариф)
    - жирные разделители после каждого пресета (Цена/Доп.услуги)
    """
    for r in range(1, rows_count + 1):
        for c in range(1, cols_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_ALL_THIN
            if r == 1:
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
                cell.fill = HEADER_FILL
            else:
                cell.alignment = CELL_ALIGN

    for c in range(1, cols_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.border = _border_with(cast(Border, cell.border), top=MEDIUM)

        cell = ws.cell(row=rows_count, column=c)
        cell.border = _border_with(cast(Border, cell.border), bottom=MEDIUM)

    for r in range(1, rows_count + 1):
        cell = ws.cell(row=r, column=1)
        cell.border = _border_with(cast(Border, cell.border), left=MEDIUM)

        cell = ws.cell(row=r, column=cols_count)
        cell.border = _border_with(cast(Border, cell.border), right=MEDIUM)

    for r in range(1, rows_count + 1):
        a = ws.cell(row=r, column=1)
        b = ws.cell(row=r, column=2)
        c = ws.cell(row=r, column=3)
        a.border = _border_with(a.border, right=MEDIUM)
        b.border = _border_with(b.border, left=MEDIUM, right=MEDIUM)
        c.border = _border_with(c.border, left=MEDIUM, right=MEDIUM)


    for i in range(1, presets_count + 1):
        end_col = 3 + i * 2
        if end_col > cols_count:
            break
        for r in range(1, rows_count + 1):
            ce = ws.cell(row=r, column=end_col)
            ce.border = _border_with(ce.border, right=MEDIUM)
            if end_col < cols_count:
                cn = ws.cell(row=r, column=end_col + 1)
                cn.border = _border_with(cn.border, left=MEDIUM)



def _apply_price_alignment(ws: Worksheet, presets_count: int) -> None:
    """Принудительно выравнивает колонки цен вправо.

    Важно: _apply_table_styles() задаёт общий CELL_ALIGN (без horizontal),
    из-за чего Excel отображает строки слева, а числа справа.
    Здесь мы фиксируем цены (в том числе строковые вида "1234 (+5%)")
    как right aligned для визуальной консистентности.
    """
    price_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        for i in range(presets_count):
            price_col = 4 + i * 2
            cell = ws.cell(row=r, column=price_col)
            cell.alignment = price_align


def _parse_days(value: Any) -> str | None:
    """Нормализует поле days для вывода в Excel.

    В исходных данных days встречается как строка диапазона (например, "1-2", "1 - 6"),
    как число/строка числа ("1"), либо пустое значение.

    Мы выводим days как текст (без приведения к int), чтобы не терять диапазоны.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Нормализуем разные тире и пробелы вокруг дефиса: "1 - 6" -> "1-6"
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s*-\s*", "-", s)
    return s



def _parse_json_dict(value: Any) -> Dict[str, Any] | None:
    """Парсит поле name_tarif.

    Поддерживает:
    - dict (уже распарсенный)
    - JSON-строку
    - строку Python-literal вида "{'A': 1, 'B': [2, '1-2']}" (часто приходит из БД)
    """
    if value is None:
        return None
    if isinstance(value, dict):
        return value
    if not isinstance(value, str):
        return None
    s = value.strip()
    if not s:
        return None

    # 1) JSON
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return obj
    except (json.JSONDecodeError, TypeError, ValueError):
        pass

    # 2) Python-literal
    try:
        obj2 = ast.literal_eval(s)
        return obj2 if isinstance(obj2, dict) else None
    except (ValueError, SyntaxError):
        return None


def _fmt_allowances(value: Any) -> str | None:
    """Функция _fmt_allowances.

    Параметры:
        value: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    d = _parse_json_dict(value)
    if not d:
        return None

    parts: List[str] = []
    for k in sorted(d.keys(), key=lambda x: str(x).lower()):
        v = d.get(k)
        if v is None:
            continue
        if isinstance(v, (int, float)):
            if isinstance(v, float) and v.is_integer():
                vs = str(int(v))
            else:
                vs = str(v).replace(".", ",")
        else:
            vs = str(v)
        parts.append(f"{k}: {vs}")

    return "; ".join(parts) if parts else None


def _fmt_num_ru(x: float) -> str:
    """Функция _fmt_num_ru.

    Параметры:
        x: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    if float(x).is_integer():
        return str(int(x))
    s = f"{x:.2f}".rstrip("0").rstrip(".")
    return s.replace(".", ",")


def _fmt_pct_ru(pct: float) -> str:
    """Функция _fmt_pct_ru.

    Параметры:
        pct: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    if float(pct).is_integer():
        s = f"{pct:+.0f}"
    else:
        s = f"{pct:+.1f}".rstrip("0").rstrip(".")
    return s.replace(".", ",") + "%"


def _price_with_pct(cur: float, prv: float) -> str:
    """Функция _price_with_pct.

    Параметры:
        cur: Описание параметра.
        prv: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    if prv == 0:
        return _fmt_num_ru(cur)
    pct = (cur - prv) / prv * 100.0
    return f"{_fmt_num_ru(cur)} ({_fmt_pct_ru(pct)})"


def _set_site_separator(ws: Worksheet, row_idx: int, cols_count: int) -> None:
    """Жирная горизонтальная линия снизу строки row_idx на всю ширину таблицы."""
    for c in range(1, cols_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = _border_with(cell.border, bottom=MEDIUM)


def _apply_empty_cells_gray(ws: Worksheet) -> None:
    """
    Серим пустые ячейки (None или ""), не трогаем шапку и не перетираем красный/зелёный.
    """
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill in (FILL_RED, FILL_GREEN):
                continue
            if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
                cell.fill = FILL_GRAY


def _normalize_url(link: str) -> str | None:
    """Функция _normalize_url.

    Параметры:
        link: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    if not link:
        return None
    url = str(link).strip()
    if not url:
        return None

    low = url.lower()


    if low.startswith("https://"):
        return url


    if url.startswith("//"):
        return "https:" + url


    return "https://" + url


def _set_site_hyperlink(ws: Worksheet, row_idx: int, site_name: str, link: str | None) -> None:
    """Функция _set_site_hyperlink.

    Параметры:
        ws: Описание параметра.
        row_idx: Описание параметра.
        site_name: Описание параметра.
        link: Описание параметра.
    """
    url = _normalize_url(link or "")
    if not url:
        return
    cell = ws.cell(row=row_idx, column=1)
    cell.value = site_name
    cell.hyperlink = url
    cell.font = HYPERLINK_FONT


def _build_sheet_for_route(
    wb: Workbook,
    route: dict,
    sites: List[dict],
    presets: List[dict],
    std_now: Dict[Tuple[int, int, int], dict],
    std_prev_price: Dict[Tuple[int, int, int], float],
    alt_tariffs: Dict[Tuple[int, int], Dict[str, Dict[int, float]]],
    alt_tariff_days: Dict[Tuple[int, int], Dict[str, str]],
    job_label: str | None = None,
) -> None:
    """Функция _build_sheet_for_route.

    Параметры:
        wb: Описание параметра.
        route: Описание параметра.
        sites: Описание параметра.
        presets: Описание параметра.
        std_now: Описание параметра.
        std_prev_price: Описание параметра.
        alt_tariffs: Описание параметра.
        job_label: Описание параметра.
    """
    title_base = f"{route['from_city']} — {route['to_city']}"
    if route.get("distance_km"):
        title_base += f", {route['distance_km']} км"
    full_title = f"{title_base} — {job_label}" if job_label else title_base
    ws = wb.create_sheet(_unique_sheet_title(wb, full_title))

    headers: List[str] = ["Компания", "Дни", "Тариф"]
    for p in presets:
        ph = _fmt_preset_header(p)
        headers.extend([f"{ph}\nЦена", f"{ph}\nДоп. услуги"])
    ws.append(headers)

    price_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    for s in sites:
        site_id = int(s["id"]) if s.get("id") is not None else s["id"]
        route_id = int(route["id"]) if route.get("id") is not None else route["id"]
        site_name = s.get("name") or ""
        site_link = s.get("link")

        site_block_start_row = ws.max_row + 1
        if site_id != 11:
            row: List[Any] = [None] * (3 + len(presets) * 2)
            row[0] = site_name
            row[2] = "Стандарт"

            first_days: str | None = None

            for i, p in enumerate(presets):
                preset_id = int(p["id"])
                k = (site_id, route_id, preset_id)
                cell_base = 3 + i * 2

                data = std_now.get(k) or {}

                if first_days is None:
                    first_days = _parse_days(data.get("days"))

                row[cell_base + 0] = data.get("price")
                row[cell_base + 1] = _fmt_allowances(data.get("allowances"))

            row[1] = first_days
            ws.append(row)
            std_row_idx = ws.max_row

            for i, p in enumerate(presets):
                preset_id = int(p["id"])
                k = (site_id, route_id, preset_id)

                price_col = 4 + i * 2
                all_col = price_col + 1

                c_price = ws.cell(row=std_row_idx, column=price_col)

                cur = std_now.get(k, {}).get("price")
                prv = std_prev_price.get(k)

                c_price.alignment = price_right

                if isinstance(cur, (int, float)) and isinstance(prv, (int, float)) and float(cur) != float(prv):
                    c_price.value = _price_with_pct(float(cur), float(prv))
                else:
                    if isinstance(c_price.value, (int, float)):
                        c_price.number_format = "#,##0.00"

                if cur is None or prv is None or (
                    isinstance(cur, (int, float)) and isinstance(prv, (int, float)) and float(cur) == float(prv)
                ):
                    c_price.fill = FILL_GRAY
                elif float(cur) < float(prv):
                    c_price.fill = FILL_GREEN
                else:
                    c_price.fill = FILL_RED

                ws.cell(row=std_row_idx, column=all_col).fill = FILL_GRAY


        tariffs_for_site_route = alt_tariffs.get((site_id, route_id)) or {}
        for tname in sorted(tariffs_for_site_route.keys(), key=lambda x: str(x).lower()):
            per_preset_prices = tariffs_for_site_route.get(tname) or {}
            trow: List[Any] = [None] * (3 + len(presets) * 2)
            trow[0] = site_name
            trow[1] = (alt_tariff_days.get((site_id, route_id)) or {}).get(str(tname))
            trow[2] = str(tname)

            for i, p in enumerate(presets):
                preset_id = int(p["id"])
                cell_base = 3 + i * 2
                trow[cell_base + 0] = per_preset_prices.get(preset_id)

            ws.append(trow)
            t_last = ws.max_row

            for i in range(len(presets)):
                price_col = 4 + i * 2
                c_price = ws.cell(row=t_last, column=price_col)
                c_price.alignment = price_right
                if isinstance(c_price.value, (int, float)):
                    c_price.number_format = "#,##0.00"
                c_price.fill = FILL_GRAY

        site_block_end_row = ws.max_row


        if site_block_end_row > site_block_start_row:
            ws.merge_cells(
                start_row=site_block_start_row,
                start_column=1,
                end_row=site_block_end_row,
                end_column=1,
            )
            top_cell = ws.cell(row=site_block_start_row, column=1)
            top_cell.alignment = Alignment(vertical="center", wrap_text=True)


        _set_site_hyperlink(ws, site_block_start_row, site_name=site_name, link=site_link)


        _set_site_separator(ws, site_block_end_row, cols_count=ws.max_column)

    _apply_table_styles(ws, ws.max_row, ws.max_column, presets_count=len(presets))
    _apply_empty_cells_gray(ws)
    _apply_price_alignment(ws, presets_count=len(presets))

    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 42
    _autosize(ws)


def export_xlsx(db: Database) -> str:
    """Функция export_xlsx.

    Параметры:
        db: Описание параметра.

    Возвращает:
        Результат выполнения функции.
    """
    sites = db.fetch_sites()
    routes = db.fetch_routes()
    presets = db.fetch_presets_ordered()

    rows = db.get_results_full_export()

    last_two = db.get_latest_done_job_ids(limit=2)
    current_job = last_two[0] if len(last_two) >= 1 else None
    prev_job = last_two[1] if len(last_two) >= 2 else None

    std_now: Dict[Tuple[int, int, int], dict] = {}
    std_prev_price: Dict[Tuple[int, int, int], float] = {}
    alt_tariffs: Dict[Tuple[int, int], Dict[str, Dict[int, float]]] = {}
    alt_tariff_days: Dict[Tuple[int, int], Dict[str, str]] = {}

    def _ingest_result(row: dict) -> None:
        """Функция _ingest_result.

        Параметры:
            r: Описание параметра.
        """
        site_id = int(row["id_site"]) if row.get("id_site") is not None else row["id_site"]
        route_id = int(row["id_route"]) if row.get("id_route") is not None else row["id_route"]
        preset_id = int(row["id_preset"]) if row.get("id_preset") is not None else row["id_preset"]
        key_triplet = (site_id, route_id, preset_id)
        std_now[key_triplet] = {
            "price": row.get("price"),
            "days": row.get("days"),
            "allowances": row.get("allowances"),
            "name_tarif": row.get("name_tarif"),
        }

        d = _parse_json_dict(row.get("name_tarif"))
        if d:
            sr = (site_id, route_id)
            bucket = alt_tariffs.setdefault(sr, {})
            for tname, tval in d.items():
                if tval is None:
                    continue

                # В некоторых сайтах (11, 13) name_tarif хранит список вида [price, days]
                days_val: str | None = None
                price_val: float | None = None

                if isinstance(tval, (int, float)):
                    price_val = float(tval)
                elif isinstance(tval, (list, tuple)) and len(tval) >= 1 and isinstance(tval[0], (int, float)):
                    price_val = float(tval[0])
                    if len(tval) >= 2:
                        days_val = _parse_days(tval[1])

                if price_val is None:
                    continue

                tkey = str(tname)
                bucket.setdefault(tkey, {})[preset_id] = price_val

                if days_val:
                    # дни сохраняем на уровень (site_id, route_id, tariff_name) — для вывода в колонку B
                    alt_tariff_days.setdefault(sr, {}).setdefault(tkey, days_val)

    if current_job is not None:
        for r in rows:
            if r.get("job_id") == current_job:
                _ingest_result(r)

        if prev_job is not None:
            for r in rows:
                if r.get("job_id") == prev_job:
                    k3 = (r["id_site"], r["id_route"], r["id_preset"])
                    pv = r.get("price")
                    if isinstance(pv, (int, float)):
                        std_prev_price[k3] = float(pv)
    else:
        by_key: Dict[Tuple[int, int, int], List[dict]] = {}
        for r in rows:
            k = (r["id_site"], r["id_route"], r["id_preset"])
            by_key.setdefault(k, []).append(r)

        for k, lst in by_key.items():
            lst.sort(key=lambda x: x.get("created_at") or "")
            last = lst[-1]
            _ingest_result(last)
            if len(lst) >= 2:
                prev = lst[-2].get("price")
                if isinstance(prev, (int, float)):
                    std_prev_price[k] = float(prev)

    job_label = f"job#{current_job}" if current_job is not None else None

    wb = Workbook()
    wb.remove(wb.active)

    for route in routes:
        _build_sheet_for_route(
            wb=wb,
            route=route,
            sites=sites,
            presets=presets,
            std_now=std_now,
            std_prev_price=std_prev_price,
            alt_tariffs=alt_tariffs,
            alt_tariff_days=alt_tariff_days,
            job_label=job_label,
        )

    export_dir = Path(__file__).resolve().parent.parent / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    out_path = export_dir / f"экспорт_{ts}.xlsx"
    wb.save(out_path)
    return str(out_path)
